import json
import requests
import openpyxl
import time
import os

file_path = os.path.dirname(__file__)+ "/coins.xlsx"
print (file_path)
wb = openpyxl.load_workbook(file_path)
sht1 = wb.worksheets[0]
lastrow=sht1.max_row
#sht1.insertrow(lastrow) and then delete all +1 of the lastrow. this way the graphs should update automatically
#there has to be a better way than inserting...
sht1.cell(row=lastrow+1,column=1).value=time.strftime("%d/%m/%Y %H:%M:%S")
i=1
j=0

def krak(ticker):
    uri = "https://api.kraken.com/0/public/Ticker"
    blah = uri + "?pair=" + ticker
    r = requests.get(blah)
    json_data = r.text
    fj = json.loads(json_data)
    print (fj)
    fuu = fj["result"][ticker]["p"][1]
    x = float(fuu)
    return x
def bittrex(ticker):
    uri = "https://bittrex.com/api/v1.1/public/"
    blah = uri + "getmarketsummary?market=" + ticker
    r = requests.get(blah)
    json_data = r.text
    fj = json.loads(json_data)
    print (fj)
    fuu = fj["result"][0]["PrevDay"]
    x = float(fuu)
    return x


for j in range(5,13):
    sht=wb.worksheets[j]
    exchange=sht.cell(row=1,column=5).value
    ticker=sht.cell(row=1,column=2).value
    coinname=sht.title
    if exchange=="Kraken":
        coinvalue=krak(ticker)
        if coinname=="Bitcoin":
            bitcoin_value=coinvalue
    elif exchange=="Bittrex":
        coinvalue=bitcoin_value*bittrex(ticker)
    sht1.cell(row=lastrow+1,column=i+1).value=coinvalue
    i+=1



wb.save("coins.xlsx")

